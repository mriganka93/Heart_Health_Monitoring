# EXPORT EXCEL
# ========================================================

from datetime import datetime
from tkinter import filedialog

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


def export_excel(
    data,
    columns
):

    filename = datetime.now().strftime(
        "health_data_%Y%m%d_%H%M%S.xlsx"
    )

    filepath = filedialog.asksaveasfilename(
        title="Export Health Data to Excel",
        defaultextension=".xlsx",
        initialfile=filename,
        filetypes=[
            (
                "Excel files",
                "*.xlsx"
            ),
            (
                "All files",
                "*.*"
            )
        ]
    )

    if not filepath:
        return False

    try:

        workbook = Workbook()

        sheet = workbook.active

        sheet.title = "Heart Health Data"

        # ====================================================
        # HEADER
        # ====================================================

        for column_index, column_name in enumerate(
            columns,
            start=1
        ):

            cell = sheet.cell(
                row=1,
                column=column_index,
                value=column_name
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.fill = PatternFill(
                start_color="0F172A",
                end_color="0F172A",
                fill_type="solid"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # ====================================================
        # DATA
        # ====================================================

        for row_index, row_data in enumerate(
            data,
            start=2
        ):

            for column_index, value in enumerate(
                row_data,
                start=1
            ):

                cell = sheet.cell(
                    row=row_index,
                    column=column_index,
                    value=value
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        # ====================================================
        # COLUMN WIDTH
        # ====================================================

        for column_cells in sheet.columns:

            max_length = 0

            column_letter = (
                column_cells[0].column_letter
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value))
                    )

            sheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 3,
                40
            )

        # ====================================================
        # FREEZE HEADER
        # ====================================================

        sheet.freeze_panes = "A2"

        # ====================================================
        # AUTO FILTER
        # ====================================================

        sheet.auto_filter.ref = sheet.dimensions

        # ====================================================
        # SAVE
        # ====================================================

        workbook.save(
            filepath
        )

        return True

    except Exception as error:

        print(
            "Excel export error:",
            error
        )

        raise
