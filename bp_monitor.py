import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import sqlite3

from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)
from reportlab.lib.styles import getSampleStyleSheet


# ============================================================
# APPLICATION CONFIGURATION
# ============================================================

APP_NAME = "Heart Health Monitor"
DB_NAME = "health_data.db"

WINDOW_WIDTH = 1200
WINDOW_HEIGHT = 850


# ============================================================
# CUSTOMTKINTER THEME
# ============================================================

ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")


# ============================================================
# COLORS
# ============================================================

COLORS = {
    "background": "#F3F6FA",
    "card": "#FFFFFF",

    "navy": "#0F172A",

    "blue": "#2563EB",
    "blue_hover": "#1D4ED8",

    "green": "#059669",
    "green_hover": "#047857",

    "red": "#DC2626",
    "red_hover": "#B91C1C",

    "gray": "#64748B",
    "light_gray": "#E2E8F0",

    "input": "#F8FAFC",

    "text": "#172033",
    "muted": "#64748B",

    "success_text": "#047857",
    "error_text": "#B91C1C",

    "table_header": "#0F172A",
    "table_even": "#FFFFFF",
    "table_odd": "#F8FAFC",
}


# ============================================================
# DATABASE
# ============================================================

def create_database():

    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    cursor.execute("""
        CREATE TABLE IF NOT EXISTS health_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            systolic INTEGER NOT NULL,
            diastolic INTEGER NOT NULL,
            heart_rate INTEGER NOT NULL DEFAULT 0,
            stress INTEGER NOT NULL DEFAULT 0,
            meditation INTEGER NOT NULL DEFAULT 0,
            heart_medication INTEGER NOT NULL DEFAULT 0,
            bp_medication INTEGER NOT NULL DEFAULT 0,
            heart_med_name TEXT DEFAULT '',
            bp_med_name TEXT DEFAULT '',
            daily_note TEXT DEFAULT '',
            exercise_duration REAL NOT NULL DEFAULT 0,
            sleep_duration REAL NOT NULL DEFAULT 0
        )
    """)

    # ========================================================
    # DATABASE MIGRATION
    # ========================================================

    cursor.execute("""
        PRAGMA table_info(health_records)
    """)

    columns = [
        row[1]
        for row in cursor.fetchall()
    ]

    migrations = {
        "heart_rate": """
            ALTER TABLE health_records
            ADD COLUMN heart_rate INTEGER NOT NULL DEFAULT 0
        """,

        "stress": """
            ALTER TABLE health_records
            ADD COLUMN stress INTEGER NOT NULL DEFAULT 0
        """,

        "meditation": """
            ALTER TABLE health_records
            ADD COLUMN meditation INTEGER NOT NULL DEFAULT 0
        """,

        "heart_medication": """
            ALTER TABLE health_records
            ADD COLUMN heart_medication INTEGER NOT NULL DEFAULT 0
        """,

        "bp_medication": """
            ALTER TABLE health_records
            ADD COLUMN bp_medication INTEGER NOT NULL DEFAULT 0
        """,

        "heart_med_name": """
            ALTER TABLE health_records
            ADD COLUMN heart_med_name TEXT DEFAULT ''
        """,

        "bp_med_name": """
            ALTER TABLE health_records
            ADD COLUMN bp_med_name TEXT DEFAULT ''
        """,

        "daily_note": """
            ALTER TABLE health_records
            ADD COLUMN daily_note TEXT DEFAULT ''
        """,

        "exercise_duration": """
            ALTER TABLE health_records
            ADD COLUMN exercise_duration REAL NOT NULL DEFAULT 0
        """,

        "sleep_duration": """
            ALTER TABLE health_records
            ADD COLUMN sleep_duration REAL NOT NULL DEFAULT 0
        """
    }

    for column_name, sql in migrations.items():

        if column_name not in columns:

            cursor.execute(sql)

    conn.commit()
    conn.close()


# ============================================================
# MAIN APPLICATION
# ============================================================

class HealthMonitor(ctk.CTk):

    def __init__(self):

        super().__init__()

        # ----------------------------------------------------
        # WINDOW
        # ----------------------------------------------------

        self.title(APP_NAME)

        self.geometry(
            f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}"
        )

        self.minsize(
            950,
            700
        )

        self.configure(
            fg_color=COLORS["background"]
        )

        try:
            self.state("zoomed")
        except Exception:
            pass

        # ====================================================
        # VARIABLES
        # ====================================================

        self.systolic_var = tk.StringVar()
        self.diastolic_var = tk.StringVar()

        self.heart_rate_var = tk.StringVar()

        self.exercise_var = tk.StringVar()
        self.sleep_var = tk.StringVar()

        self.stress_var = tk.BooleanVar(
            value=False
        )

        self.meditation_var = tk.BooleanVar(
            value=False
        )

        # ====================================================
        # MEDICATION VARIABLES
        # ====================================================

        self.heart_medication_var = tk.BooleanVar(
            value=False
        )

        self.bp_medication_var = tk.BooleanVar(
            value=False
        )

        self.heart_med_name_var = tk.StringVar()
        self.bp_med_name_var = tk.StringVar()

        # ====================================================
        # EXPORT VARIABLES
        # ====================================================

        self.export_format_var = tk.StringVar(
            value="PDF"
        )

        self.export_columns = [
            "Date / Time",
            "Systolic Pressure",
            "Diastolic Pressure",
            "Heart Rate",
            "Stress",
            "Meditation",
            "Heart Medication",
            "Heart Medication Name",
            "BP Medication",
            "BP Medication Name",
            "Exercise Duration",
            "Sleep Duration",
            "Daily Note"
        ]

        # ====================================================
        # BUILD UI
        # ====================================================

        self.create_header()

        self.create_scroll_area()

        self.create_health_entry_card()

        self.create_recent_data_card()

        # ====================================================
        # KEYBOARD SHORTCUTS
        # ====================================================

        self.bind(
            "<Control-s>",
            lambda event: self.save_record()
        )

        self.bind(
            "<Escape>",
            lambda event: self.clear_form()
        )

    # ========================================================
    # HEADER
    # ========================================================

    def create_header(self):

        self.header = ctk.CTkFrame(
            self,
            height=95,
            corner_radius=0,
            fg_color=COLORS["navy"]
        )

        self.header.pack(
            fill="x"
        )

        self.header.pack_propagate(False)

        header_inner = ctk.CTkFrame(
            self.header,
            fg_color="transparent"
        )

        header_inner.pack(
            fill="both",
            expand=True,
            padx=45
        )

        # ----------------------------------------------------
        # LOGO
        # ----------------------------------------------------

        logo = ctk.CTkFrame(
            header_inner,
            width=50,
            height=50,
            corner_radius=13,
            fg_color=COLORS["blue"]
        )

        logo.pack(
            side="left",
            pady=22
        )

        logo.pack_propagate(False)

        logo_text = ctk.CTkLabel(
            logo,
            text="♥",
            text_color="white",
            font=(
                "Arial",
                25,
                "bold"
            )
        )

        logo_text.pack(
            expand=True
        )

        # ----------------------------------------------------
        # TITLE
        # ----------------------------------------------------

        title_frame = ctk.CTkFrame(
            header_inner,
            fg_color="transparent"
        )

        title_frame.pack(
            side="left",
            padx=17
        )

        title = ctk.CTkLabel(
            title_frame,
            text="Heart Health Monitor",
            text_color="white",
            font=(
                "Segoe UI",
                30,
                "bold"
            )
        )

        title.pack(
            anchor="w"
        )

        subtitle = ctk.CTkLabel(
            title_frame,
            text="Personal Heart Health Tracking Developed and Build By @Mriganka93",
            text_color="#94A3B8",
            font=(
                "Segoe UI",
                13
            )
        )

        subtitle.pack(
            anchor="w"
        )

        # ----------------------------------------------------
        # DATE
        # ----------------------------------------------------

        today = datetime.now().strftime(
            "%A, %d %B %Y"
        )

        date_label = ctk.CTkLabel(
            header_inner,
            text=today,
            text_color="#CBD5E1",
            font=(
                "Segoe UI",
                13
            )
        )

        date_label.pack(
            side="right"
        )

    # ========================================================
    # SCROLLABLE MAIN AREA
    # ========================================================

    def create_scroll_area(self):

        self.scroll_area = ctk.CTkScrollableFrame(
            self,
            fg_color=COLORS["background"],
            scrollbar_button_color="#CBD5E1",
            scrollbar_button_hover_color="#94A3B8"
        )

        self.scroll_area.pack(
            fill="both",
            expand=True
        )

        self.scroll_area.grid_columnconfigure(
            0,
            weight=1
        )

    # ========================================================
    # HEALTH ENTRY CARD
    # ========================================================

    def create_health_entry_card(self):

        card = ctk.CTkFrame(
            self.scroll_area,
            fg_color=COLORS["card"],
            corner_radius=18,
            border_width=1,
            border_color=COLORS["light_gray"]
        )

        card.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=65,
            pady=(35, 20)
        )

        card.grid_columnconfigure(
            0,
            weight=1
        )

        # ----------------------------------------------------
        # HEADING
        # ----------------------------------------------------

        heading = ctk.CTkFrame(
            card,
            fg_color="transparent"
        )

        heading.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=35,
            pady=(30, 5)
        )

        title = ctk.CTkLabel(
            heading,
            text="Daily Health Entry",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                25,
                "bold"
            )
        )

        title.pack(
            side="left"
        )

        badge = ctk.CTkLabel(
            heading,
            text="  TODAY  ",
            text_color="#1D4ED8",
            fg_color="#DBEAFE",
            corner_radius=6,
            font=(
                "Segoe UI",
                11,
                "bold"
            )
        )

        badge.pack(
            side="left",
            padx=14
        )

        # ----------------------------------------------------
        # DESCRIPTION
        # ----------------------------------------------------

        description = ctk.CTkLabel(
            card,
            text=(
                "Record your blood pressure, heart rate, "
                "activity, sleep and daily wellness."
            ),
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                15,
                "bold"
            )
        )

        description.grid(
            row=1,
            column=0,
            sticky="w",
            padx=35,
            pady=(0, 25)
        )

        # ----------------------------------------------------
        # INPUT GRID
        # ----------------------------------------------------

        inputs = ctk.CTkFrame(
            card,
            fg_color="transparent"
        )

        inputs.grid(
            row=2,
            column=0,
            sticky="ew",
            padx=28
        )

        inputs.grid_columnconfigure(
            0,
            weight=1
        )

        inputs.grid_columnconfigure(
            1,
            weight=1
        )

        # ----------------------------------------------------
        # BLOOD PRESSURE
        # ----------------------------------------------------

        self.create_input(
            inputs,
            "Systolic Pressure",
            "mmHg",
            "e.g. 120",
            self.systolic_var,
            0,
            0
        )

        self.create_input(
            inputs,
            "Diastolic Pressure",
            "mmHg",
            "e.g. 80",
            self.diastolic_var,
            0,
            1
        )

        # ----------------------------------------------------
        # HEART RATE
        # ----------------------------------------------------

        self.create_input(
            inputs,
            "Heart Rate",
            "BPM",
            "e.g. 72",
            self.heart_rate_var,
            1,
            0
        )

        # ----------------------------------------------------
        # EXERCISE
        # ----------------------------------------------------

        self.create_input(
            inputs,
            "Exercise Duration",
            "minutes",
            "e.g. 30",
            self.exercise_var,
            1,
            1
        )

        # ----------------------------------------------------
        # SLEEP
        # ----------------------------------------------------

        self.create_input(
            inputs,
            "Sleep Duration",
            "hours",
            "e.g. 7.5",
            self.sleep_var,
            2,
            0
        )

        # ====================================================
        # WELLNESS
        # ====================================================

        wellness = ctk.CTkFrame(
            inputs,
            fg_color="#F8FAFC",
            corner_radius=12,
            border_width=1,
            border_color=COLORS["light_gray"]
        )

        wellness.grid(
            row=3,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=7,
            pady=(20, 8)
        )

        wellness.grid_columnconfigure(
            0,
            weight=1
        )

        wellness.grid_columnconfigure(
            1,
            weight=1
        )

        wellness_title = ctk.CTkLabel(
            wellness,
            text="Daily Wellness",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                15,
                "bold"
            )
        )

        wellness_title.grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            padx=18,
            pady=(15, 10)
        )

        # ----------------------------------------------------
        # STRESS
        # ----------------------------------------------------

        stress_box = ctk.CTkCheckBox(
            wellness,
            text="Stress",
            variable=self.stress_var,
            text_color=COLORS["text"],
            hover_color=COLORS["blue_hover"],
            fg_color=COLORS["blue"],
            border_color="#94A3B8",
            corner_radius=5,
            checkbox_width=24,
            checkbox_height=24,
            font=(
                "Segoe UI",
                14
            )
        )

        stress_box.grid(
            row=1,
            column=0,
            sticky="w",
            padx=18,
            pady=(2, 19)
        )

        # ----------------------------------------------------
        # MEDITATION
        # ----------------------------------------------------

        meditation_box = ctk.CTkCheckBox(
            wellness,
            text="Meditation",
            variable=self.meditation_var,
            text_color=COLORS["text"],
            hover_color=COLORS["blue_hover"],
            fg_color=COLORS["blue"],
            border_color="#94A3B8",
            corner_radius=5,
            checkbox_width=24,
            checkbox_height=24,
            font=(
                "Segoe UI",
                14
            )
        )

        meditation_box.grid(
            row=1,
            column=1,
            sticky="w",
            padx=18,
            pady=(2, 19)
        )

        # ====================================================
        # MEDICATION
        # ====================================================

        medication = ctk.CTkFrame(
            inputs,
            fg_color="#EFF6FF",
            corner_radius=12,
            border_width=1,
            border_color="#BFDBFE"
        )

        medication.grid(
            row=4,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=7,
            pady=(10, 8)
        )

        medication.grid_columnconfigure(
            0,
            weight=1
        )

        medication.grid_columnconfigure(
            1,
            weight=1
        )

        medication_title = ctk.CTkLabel(
            medication,
            text="Medication",
            text_color="#1E40AF",
            font=(
                "Segoe UI",
                15,
                "bold"
            )
        )

        medication_title.grid(
            row=0,
            column=0,
            columnspan=2,
            sticky="w",
            padx=18,
            pady=(15, 10)
        )

        # ----------------------------------------------------
        # HEART MEDICATION CHECKBOX
        # ----------------------------------------------------

        heart_med_box = ctk.CTkCheckBox(
            medication,
            text="Heart Meds",
            variable=self.heart_medication_var,
            text_color=COLORS["text"],
            hover_color=COLORS["blue_hover"],
            fg_color=COLORS["blue"],
            border_color="#94A3B8",
            corner_radius=5,
            checkbox_width=24,
            checkbox_height=24,
            font=(
                "Segoe UI",
                14
            )
        )

        heart_med_box.grid(
            row=1,
            column=0,
            sticky="w",
            padx=18,
            pady=(2, 8)
        )

        # ----------------------------------------------------
        # BP MEDICATION CHECKBOX
        # ----------------------------------------------------

        bp_med_box = ctk.CTkCheckBox(
            medication,
            text="BP Meds",
            variable=self.bp_medication_var,
            text_color=COLORS["text"],
            hover_color=COLORS["blue_hover"],
            fg_color=COLORS["blue"],
            border_color="#94A3B8",
            corner_radius=5,
            checkbox_width=24,
            checkbox_height=24,
            font=(
                "Segoe UI",
                14
            )
        )

        bp_med_box.grid(
            row=1,
            column=1,
            sticky="w",
            padx=18,
            pady=(2, 8)
        )

        # ----------------------------------------------------
        # HEART MED NAME
        # ----------------------------------------------------

        heart_name_label = ctk.CTkLabel(
            medication,
            text="Heart Med Name (optional)",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                12,
                "bold"
            )
        )

        heart_name_label.grid(
            row=2,
            column=0,
            sticky="w",
            padx=18,
            pady=(5, 5)
        )

        heart_name_entry = ctk.CTkEntry(
            medication,
            textvariable=self.heart_med_name_var,
            height=42,
            corner_radius=8,
            border_width=1,
            border_color=COLORS["light_gray"],
            fg_color="white",
            text_color=COLORS["text"],
            placeholder_text="e.g. Metoprolol",
            placeholder_text_color="#94A3B8",
            font=(
                "Segoe UI",
                13
            )
        )

        heart_name_entry.grid(
            row=3,
            column=0,
            sticky="ew",
            padx=18,
            pady=(0, 15)
        )

        # ----------------------------------------------------
        # BP MED NAME
        # ----------------------------------------------------

        bp_name_label = ctk.CTkLabel(
            medication,
            text="BP Med Name (optional)",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                12,
                "bold"
            )
        )

        bp_name_label.grid(
            row=2,
            column=1,
            sticky="w",
            padx=18,
            pady=(5, 5)
        )

        bp_name_entry = ctk.CTkEntry(
            medication,
            textvariable=self.bp_med_name_var,
            height=42,
            corner_radius=8,
            border_width=1,
            border_color=COLORS["light_gray"],
            fg_color="white",
            text_color=COLORS["text"],
            placeholder_text="e.g. Amlodipine",
            placeholder_text_color="#94A3B8",
            font=(
                "Segoe UI",
                13
            )
        )

        bp_name_entry.grid(
            row=3,
            column=1,
            sticky="ew",
            padx=18,
            pady=(0, 15)
        )

        # ====================================================
        # DAILY NOTE
        # ====================================================

        note_frame = ctk.CTkFrame(
            inputs,
            fg_color="#F8FAFC",
            corner_radius=12,
            border_width=1,
            border_color=COLORS["light_gray"]
        )

        note_frame.grid(
            row=5,
            column=0,
            columnspan=2,
            sticky="ew",
            padx=7,
            pady=(10, 8)
        )

        note_frame.grid_columnconfigure(
            0,
            weight=1
        )

        note_title = ctk.CTkLabel(
            note_frame,
            text="Daily Note",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                15,
                "bold"
            )
        )

        note_title.grid(
            row=0,
            column=0,
            sticky="w",
            padx=18,
            pady=(15, 2)
        )

        note_description = ctk.CTkLabel(
            note_frame,
            text="How did you feel today? Add any notes you want to remember.",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                12
            )
        )

        note_description.grid(
            row=1,
            column=0,
            sticky="w",
            padx=18,
            pady=(0, 8)
        )

        self.daily_note_text = ctk.CTkTextbox(
            note_frame,
            height=100,
            corner_radius=8,
            border_width=1,
            border_color=COLORS["light_gray"],
            fg_color="white",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                13
            )
        )

        self.daily_note_text.grid(
            row=2,
            column=0,
            sticky="ew",
            padx=18,
            pady=(0, 18)
        )

        # ====================================================
        # BUTTON BAR
        # ====================================================

        buttons = ctk.CTkFrame(
            card,
            fg_color="transparent"
        )

        buttons.grid(
            row=3,
            column=0,
            sticky="ew",
            padx=35,
            pady=(22, 10)
        )

        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        save_button = ctk.CTkButton(
            buttons,
            text="Save Health Data",
            command=self.save_record,
            width=180,
            height=46,
            corner_radius=9,
            fg_color=COLORS["blue"],
            hover_color=COLORS["blue_hover"],
            font=(
                "Segoe UI",
                14,
                "bold"
            )
        )

        save_button.pack(
            side="right"
        )

        # ----------------------------------------------------
        # CLEAR
        # ----------------------------------------------------

        clear_button = ctk.CTkButton(
            buttons,
            text="Clear",
            command=self.clear_form,
            width=105,
            height=46,
            corner_radius=9,
            fg_color="#64748B",
            hover_color="#475569",
            font=(
                "Segoe UI",
                14,
                "bold"
            )
        )

        clear_button.pack(
            side="right",
            padx=10
        )

        # ----------------------------------------------------
        # DELETE
        # ----------------------------------------------------

        delete_button = ctk.CTkButton(
            buttons,
            text="Delete Last",
            command=self.delete_last_record,
            width=135,
            height=46,
            corner_radius=9,
            fg_color=COLORS["red"],
            hover_color=COLORS["red_hover"],
            font=(
                "Segoe UI",
                14,
                "bold"
            )
        )

        delete_button.pack(
            side="right"
        )

        # ----------------------------------------------------
        # STATUS
        # ----------------------------------------------------

        self.status_frame = ctk.CTkFrame(
            card,
            fg_color="transparent"
        )

        self.status_frame.grid(
            row=4,
            column=0,
            sticky="ew",
            padx=35,
            pady=(0, 25)
        )

        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="Ready",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                13
            )
        )

        self.status_label.pack(
            anchor="w"
        )

    # ========================================================
    # INPUT FIELD
    # ========================================================

    def create_input(
        self,
        parent,
        label,
        unit,
        placeholder,
        variable,
        row,
        column
    ):

        frame = ctk.CTkFrame(
            parent,
            fg_color="transparent"
        )

        frame.grid(
            row=row,
            column=column,
            sticky="ew",
            padx=7,
            pady=9
        )

        frame.grid_columnconfigure(
            0,
            weight=1
        )

        label_frame = ctk.CTkFrame(
            frame,
            fg_color="transparent"
        )

        label_frame.grid(
            row=0,
            column=0,
            sticky="ew",
            pady=(0, 7)
        )

        label_widget = ctk.CTkLabel(
            label_frame,
            text=label,
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                14,
                "bold"
            )
        )

        label_widget.pack(
            side="left"
        )

        unit_widget = ctk.CTkLabel(
            label_frame,
            text=unit,
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                12
            )
        )

        unit_widget.pack(
            side="right"
        )

        entry = ctk.CTkEntry(
            frame,
            textvariable=variable,
            height=50,
            corner_radius=9,
            border_width=1,
            border_color=COLORS["light_gray"],
            fg_color=COLORS["input"],
            text_color=COLORS["text"],
            placeholder_text=placeholder,
            placeholder_text_color="#94A3B8",
            font=(
                "Segoe UI",
                15
            )
        )

        entry.grid(
            row=1,
            column=0,
            sticky="ew"
        )

    # ========================================================
    # RECENT DATA CARD
    # ========================================================

    def create_recent_data_card(self):

        card = ctk.CTkFrame(
            self.scroll_area,
            fg_color=COLORS["card"],
            corner_radius=18,
            border_width=1,
            border_color=COLORS["light_gray"]
        )

        card.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=65,
            pady=(0, 40)
        )

        card.grid_columnconfigure(
            0,
            weight=1
        )

        # ----------------------------------------------------
        # HEADER
        # ----------------------------------------------------

        header = ctk.CTkFrame(
            card,
            fg_color="transparent"
        )

        header.grid(
            row=0,
            column=0,
            sticky="ew",
            padx=35,
            pady=(30, 20)
        )

        title = ctk.CTkLabel(
            header,
            text="Recent Health Data",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                25,
                "bold"
            )
        )

        title.pack(
            side="left"
        )

        subtitle = ctk.CTkLabel(
            header,
            text="  LAST 5 DAYS  ",
            text_color="#475569",
            fg_color="#F1F5F9",
            corner_radius=6,
            font=(
                "Segoe UI",
                11,
                "bold"
            )
        )

        subtitle.pack(
            side="left",
            padx=14
        )

        # ----------------------------------------------------
        # VIEW DATA
        # ----------------------------------------------------

        view_button = ctk.CTkButton(
            header,
            text="View Data",
            command=self.view_data,
            width=135,
            height=42,
            corner_radius=8,
            fg_color=COLORS["blue"],
            hover_color=COLORS["blue_hover"],
            font=(
                "Segoe UI",
                13,
                "bold"
            )
        )

        view_button.pack(
            side="right"
        )

        # ====================================================
        # TABLE
        # ====================================================

        table_container = ctk.CTkFrame(
            card,
            fg_color="white",
            corner_radius=10,
            border_width=1,
            border_color=COLORS["light_gray"]
        )

        table_container.grid(
            row=1,
            column=0,
            sticky="ew",
            padx=35
        )

        style = ttk.Style()

        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(
            "Health.Treeview",
            background="white",
            foreground=COLORS["text"],
            fieldbackground="white",
            rowheight=50,
            borderwidth=0,
            font=(
                "Segoe UI",
                12
            )
        )

        style.configure(
            "Health.Treeview.Heading",
            background=COLORS["table_header"],
            foreground="white",
            font=(
                "Segoe UI",
                12,
                "bold"
            ),
            relief="flat",
            padding=12
        )

        style.map(
            "Health.Treeview",
            background=[
                (
                    "selected",
                    "#DBEAFE"
                )
            ],
            foreground=[
                (
                    "selected",
                    COLORS["text"]
                )
            ]
        )

        columns = (
            "date",
            "systolic",
            "diastolic",
            "heart_rate",
            "stress",
            "meditation",
            "heart_med",
            "bp_med",
            "exercise",
            "sleep"
        )

        self.table = ttk.Treeview(
            table_container,
            columns=columns,
            show="headings",
            height=7,
            style="Health.Treeview"
        )

        headers = {
            "date": "Date / Time",
            "systolic": "Systolic",
            "diastolic": "Diastolic",
            "heart_rate": "Heart Rate",
            "stress": "Stress",
            "meditation": "Meditation",
            "heart_med": "Heart Med",
            "bp_med": "BP Med",
            "exercise": "Exercise",
            "sleep": "Sleep"
        }

        widths = {
            "date": 180,
            "systolic": 100,
            "diastolic": 100,
            "heart_rate": 115,
            "stress": 80,
            "meditation": 105,
            "heart_med": 110,
            "bp_med": 100,
            "exercise": 110,
            "sleep": 100
        }

        for column in columns:

            self.table.heading(
                column,
                text=headers[column]
            )

            self.table.column(
                column,
                width=widths[column],
                minwidth=75,
                anchor="center"
            )

        self.table.pack(
            fill="x",
            expand=True
        )

        self.table.tag_configure(
            "even",
            background=COLORS["table_even"]
        )

        self.table.tag_configure(
            "odd",
            background=COLORS["table_odd"]
        )

        self.table_status = ctk.CTkLabel(
            card,
            text=(
                "Click View Data to display "
                "your recent health records."
            ),
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                13
            )
        )

        self.table_status.grid(
            row=2,
            column=0,
            sticky="w",
            padx=35,
            pady=(12, 0)
        )

        # ====================================================
        # EXPORT SECTION
        # ====================================================

        export_frame = ctk.CTkFrame(
            card,
            fg_color="#F8FAFC",
            corner_radius=12
        )

        export_frame.grid(
            row=3,
            column=0,
            sticky="ew",
            padx=35,
            pady=(20, 30)
        )

        export_info = ctk.CTkLabel(
            export_frame,
            text="Export selected health data:",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                13
            )
        )

        export_info.grid(
           row=0,
           column=0,
           sticky="w",
           padx=(300, 20),
           pady=15
        )

        # ----------------------------------------------------
        # FORMAT
        # ----------------------------------------------------

        format_label = ctk.CTkLabel(
            export_frame,
            text="Format:",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                13,
                "bold"
            )
        )

        format_label.grid(
            row=0,
            column=1,
            padx=(10, 5),
            pady=10
        )

        format_menu = ctk.CTkOptionMenu(
            export_frame,
            variable=self.export_format_var,
            values=[
                "PDF",
                "Excel"
            ],
            width=110,
            height=40,
            corner_radius=8,
            fg_color=COLORS["blue"],
            button_color=COLORS["blue"],
            button_hover_color=COLORS["blue_hover"],
            dropdown_fg_color="white",
            dropdown_hover_color="#DBEAFE",
            text_color="white",
            font=(
                "Segoe UI",
                13,
                "bold"
            )
        )

        format_menu.grid(
            row=0,
            column=2,
            padx=5,
            pady=10
        )

        # ----------------------------------------------------
        # SELECT COLUMNS
        # ----------------------------------------------------

        select_columns_button = ctk.CTkButton(
            export_frame,
            text="Select Data Columns",
            command=self.select_export_columns,
            width=175,
            height=40,
            corner_radius=8,
            fg_color="#64748B",
            hover_color="#475569",
            font=(
                "Segoe UI",
                13,
                "bold"
            )
        )

        select_columns_button.grid(
            row=0,
            column=3,
            padx=5,
            pady=10
        )

        # ----------------------------------------------------
        # EXPORT
        # ----------------------------------------------------

        export_button = ctk.CTkButton(
            export_frame,
            text="Export",
            command=self.export_data,
            width=120,
            height=40,
            corner_radius=8,
            fg_color=COLORS["green"],
            hover_color=COLORS["green_hover"],
            font=(
                "Segoe UI",
                13,
                "bold"
            )
        )

        export_button.grid(
            row=0,
            column=4,
            padx=(5, 12),
            pady=10
        )

    # ========================================================
    # SAVE RECORD
    # ========================================================

    def save_record(self):

        # ----------------------------------------------------
        # SYSTOLIC
        # ----------------------------------------------------

        try:
            systolic = int(
                self.systolic_var.get()
            )
        except ValueError:

            self.show_error(
                "Please enter a valid systolic pressure."
            )

            return

        # ----------------------------------------------------
        # DIASTOLIC
        # ----------------------------------------------------

        try:
            diastolic = int(
                self.diastolic_var.get()
            )
        except ValueError:

            self.show_error(
                "Please enter a valid diastolic pressure."
            )

            return

        # ----------------------------------------------------
        # HEART RATE
        # ----------------------------------------------------

        try:
            heart_rate = int(
                self.heart_rate_var.get()
            )
        except ValueError:

            self.show_error(
                "Please enter a valid heart rate."
            )

            return

        # ----------------------------------------------------
        # EXERCISE
        # ----------------------------------------------------

        try:
            exercise = float(
                self.exercise_var.get()
            )
        except ValueError:

            self.show_error(
                "Please enter exercise duration in minutes."
            )

            return

        # ----------------------------------------------------
        # SLEEP
        # ----------------------------------------------------

        try:
            sleep = float(
                self.sleep_var.get()
            )
        except ValueError:

            self.show_error(
                "Please enter sleep duration in hours."
            )

            return

        # ----------------------------------------------------
        # VALIDATION
        # ----------------------------------------------------

        if systolic <= 0:

            self.show_error(
                "Systolic pressure must be greater than zero."
            )

            return

        if diastolic <= 0:

            self.show_error(
                "Diastolic pressure must be greater than zero."
            )

            return

        if heart_rate <= 0:

            self.show_error(
                "Heart rate must be greater than zero."
            )

            return

        if exercise < 0:

            self.show_error(
                "Exercise duration cannot be negative."
            )

            return

        if sleep < 0:

            self.show_error(
                "Sleep duration cannot be negative."
            )

            return

        if sleep > 24:

            self.show_error(
                "Sleep duration cannot exceed 24 hours."
            )

            return

        # ----------------------------------------------------
        # FLAGS
        # ----------------------------------------------------

        stress = int(
            self.stress_var.get()
        )

        meditation = int(
            self.meditation_var.get()
        )

        heart_medication = int(
            self.heart_medication_var.get()
        )

        bp_medication = int(
            self.bp_medication_var.get()
        )

        # ----------------------------------------------------
        # OPTIONAL MEDICATION NAMES
        # ----------------------------------------------------

        heart_med_name = (
            self.heart_med_name_var.get().strip()
        )

        bp_med_name = (
            self.bp_med_name_var.get().strip()
        )

        # ----------------------------------------------------
        # DAILY NOTE
        # ----------------------------------------------------

        daily_note = (
            self.daily_note_text.get(
                "1.0",
                "end"
            ).strip()
        )

        # ----------------------------------------------------
        # SAVE
        # ----------------------------------------------------

        try:

            conn = sqlite3.connect(
                DB_NAME
            )

            cursor = conn.cursor()

            timestamp = datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            )

            cursor.execute("""
                INSERT INTO health_records (
                    timestamp,
                    systolic,
                    diastolic,
                    heart_rate,
                    stress,
                    meditation,
                    heart_medication,
                    bp_medication,
                    heart_med_name,
                    bp_med_name,
                    daily_note,
                    exercise_duration,
                    sleep_duration
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                timestamp,
                systolic,
                diastolic,
                heart_rate,
                stress,
                meditation,
                heart_medication,
                bp_medication,
                heart_med_name,
                bp_med_name,
                daily_note,
                exercise,
                sleep
            ))

            conn.commit()
            conn.close()

        except sqlite3.Error as error:

            print(
                "Database error:",
                error
            )

            self.show_error(
                "Unable to save data to the database."
            )

            return

        # ----------------------------------------------------
        # SUCCESS
        # ----------------------------------------------------

        self.clear_form(
            update_status=False
        )

        self.show_success(
            "Health data saved successfully."
        )

    # ========================================================
    # VIEW DATA
    # ========================================================

    def view_data(self):

        conn = sqlite3.connect(
            DB_NAME
        )

        cursor = conn.cursor()

        start_date = (
            datetime.now().date()
            - timedelta(days=4)
        )

        start_datetime = datetime.combine(
            start_date,
            datetime.min.time()
        ).strftime(
            "%Y-%m-%d %H:%M:%S"
        )

        cursor.execute("""
            SELECT
                timestamp,
                systolic,
                diastolic,
                heart_rate,
                stress,
                meditation,
                heart_medication,
                bp_medication,
                exercise_duration,
                sleep_duration
            FROM health_records
            WHERE timestamp >= ?
            ORDER BY timestamp DESC
        """, (
            start_datetime,
        ))

        records = cursor.fetchall()

        conn.close()

        for item in self.table.get_children():

            self.table.delete(
                item
            )

        for index, record in enumerate(records):

            (
                timestamp,
                systolic,
                diastolic,
                heart_rate,
                stress,
                meditation,
                heart_medication,
                bp_medication,
                exercise,
                sleep
            ) = record

            try:

                date_obj = datetime.strptime(
                    timestamp,
                    "%Y-%m-%d %H:%M:%S"
                )

                display_date = date_obj.strftime(
                    "%d %b %Y  %I:%M %p"
                )

            except ValueError:

                display_date = timestamp

            stress_text = (
                "Yes" if stress else "No"
            )

            meditation_text = (
                "Yes" if meditation else "No"
            )

            heart_med_text = (
                "Yes" if heart_medication else "No"
            )

            bp_med_text = (
                "Yes" if bp_medication else "No"
            )

            row_tag = (
                "even"
                if index % 2 == 0
                else "odd"
            )

            self.table.insert(
                "",
                "end",
                values=(
                    display_date,
                    systolic,
                    diastolic,
                    f"{heart_rate} BPM",
                    stress_text,
                    meditation_text,
                    heart_med_text,
                    bp_med_text,
                    f"{exercise:g} min",
                    f"{sleep:g} hr"
                ),
                tags=(row_tag,)
            )

        self.table_status.configure(
            text=(
                f"{len(records)} record(s) "
                "found in the last 5 days."
            )
        )

    # ========================================================
    # DELETE LAST RECORD
    # ========================================================

    def delete_last_record(self):

        conn = sqlite3.connect(
            DB_NAME
        )

        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                id,
                timestamp,
                systolic,
                diastolic,
                heart_rate,
                stress,
                meditation,
                heart_medication,
                bp_medication,
                heart_med_name,
                bp_med_name,
                daily_note,
                exercise_duration,
                sleep_duration
            FROM health_records
            ORDER BY id DESC
            LIMIT 1
        """)

        record = cursor.fetchone()

        if record is None:

            conn.close()

            self.show_error(
                "There are no saved records to delete."
            )

            return

        (
            record_id,
            timestamp,
            systolic,
            diastolic,
            heart_rate,
            stress,
            meditation,
            heart_medication,
            bp_medication,
            heart_med_name,
            bp_med_name,
            daily_note,
            exercise,
            sleep
        ) = record

        answer = messagebox.askyesno(
            "Delete Last Record",

            (
                "Are you sure you want to delete "
                "the last saved record?\n\n"
                f"Date: {timestamp}\n"
                f"Systolic: {systolic} mmHg\n"
                f"Diastolic: {diastolic} mmHg\n"
                f"Heart Rate: {heart_rate} BPM\n"
                f"Stress: {'Yes' if stress else 'No'}\n"
                f"Meditation: {'Yes' if meditation else 'No'}\n"
                f"Heart Medication: "
                f"{'Yes' if heart_medication else 'No'}\n"
                f"Heart Med Name: "
                f"{heart_med_name or 'Not entered'}\n"
                f"BP Medication: "
                f"{'Yes' if bp_medication else 'No'}\n"
                f"BP Med Name: "
                f"{bp_med_name or 'Not entered'}\n"
                f"Exercise: {exercise:g} minutes\n"
                f"Sleep: {sleep:g} hours"
            ),

            icon="warning"
        )

        if not answer:

            conn.close()

            return

        cursor.execute("""
            DELETE FROM health_records
            WHERE id = ?
        """, (
            record_id,
        ))

        conn.commit()
        conn.close()

        self.show_success(
            "Last saved record deleted."
        )

    # ========================================================
    # SELECT EXPORT COLUMNS
    # ========================================================

    def select_export_columns(self):

        window = ctk.CTkToplevel(
            self
        )

        window.title(
            "Select Data Columns"
        )

        window.geometry(
            "520x720"
        )

        window.resizable(
            False,
            False
        )

        window.transient(
            self
        )

        window.grab_set()

        # ----------------------------------------------------
        # TITLE
        # ----------------------------------------------------

        title = ctk.CTkLabel(
            window,
            text="Select Data Columns",
            text_color=COLORS["text"],
            font=(
                "Segoe UI",
                22,
                "bold"
            )
        )

        title.pack(
            pady=(25, 5)
        )

        description = ctk.CTkLabel(
            window,
            text="Choose the information you want to export.",
            text_color=COLORS["muted"],
            font=(
                "Segoe UI",
                13
            )
        )

        description.pack(
            pady=(0, 20)
        )

        # ----------------------------------------------------
        # CHECKBOX CONTAINER
        # ----------------------------------------------------

        checkbox_frame = ctk.CTkScrollableFrame(
            window,
            fg_color="#F8FAFC",
            corner_radius=12
        )

        checkbox_frame.pack(
            fill="both",
            expand=True,
            padx=25,
            pady=5
        )

        # ----------------------------------------------------
        # ALL AVAILABLE COLUMNS
        # ----------------------------------------------------

        available_columns = [
            "Date / Time",
            "Systolic Pressure",
            "Diastolic Pressure",
            "Heart Rate",
            "Stress",
            "Meditation",
            "Heart Medication",
            "Heart Medication Name",
            "BP Medication",
            "BP Medication Name",
            "Exercise Duration",
            "Sleep Duration",
            "Daily Note"
        ]

        column_vars = {}

        for column in available_columns:

            variable = tk.BooleanVar(
                value=column in self.export_columns
            )

            column_vars[column] = variable

            checkbox = ctk.CTkCheckBox(
                checkbox_frame,
                text=column,
                variable=variable,
                text_color=COLORS["text"],
                fg_color=COLORS["blue"],
                hover_color=COLORS["blue_hover"],
                border_color="#94A3B8",
                checkbox_width=22,
                checkbox_height=22,
                font=(
                    "Segoe UI",
                    14
                )
            )

            checkbox.pack(
                anchor="w",
                padx=20,
                pady=7
            )

        # ----------------------------------------------------
        # BUTTON FRAME
        # ----------------------------------------------------

        button_frame = ctk.CTkFrame(
            window,
            fg_color="transparent"
        )

        button_frame.pack(
            fill="x",
            padx=25,
            pady=20
        )

        # ----------------------------------------------------
        # SELECT ALL
        # ----------------------------------------------------

        def select_all():

            for variable in column_vars.values():

                variable.set(True)

        select_all_button = ctk.CTkButton(
            button_frame,
            text="Select All",
            command=select_all,
            width=110,
            height=40,
            corner_radius=8,
            fg_color="#64748B",
            hover_color="#475569",
            font=(
                "Segoe UI",
                12,
                "bold"
            )
        )

        select_all_button.pack(
            side="left"
        )

        # ----------------------------------------------------
        # CLEAR ALL
        # ----------------------------------------------------

        def clear_all():

            for variable in column_vars.values():

                variable.set(False)

        clear_all_button = ctk.CTkButton(
            button_frame,
            text="Clear All",
            command=clear_all,
            width=110,
            height=40,
            corner_radius=8,
            fg_color="#64748B",
            hover_color="#475569",
            font=(
                "Segoe UI",
                12,
                "bold"
            )
        )

        clear_all_button.pack(
            side="left",
            padx=8
        )

        # ----------------------------------------------------
        # APPLY
        # ----------------------------------------------------

        def apply_selection():

            selected = [
                column
                for column, variable in column_vars.items()
                if variable.get()
            ]

            if not selected:

                messagebox.showwarning(
                    "No Columns Selected",
                    "Please select at least one data column.",
                    parent=window
                )

                return

            self.export_columns = selected

            window.destroy()

            self.status_label.configure(
                text=(
                    f"✓  {len(selected)} export column(s) selected."
                ),
                text_color=COLORS["success_text"]
            )

        apply_button = ctk.CTkButton(
            button_frame,
            text="Apply",
            command=apply_selection,
            width=110,
            height=40,
            corner_radius=8,
            fg_color=COLORS["blue"],
            hover_color=COLORS["blue_hover"],
            font=(
                "Segoe UI",
                12,
                "bold"
            )
        )

        apply_button.pack(
            side="right"
        )

    # ========================================================
    # EXPORT DATA
    # ========================================================

    def export_data(self):

        if not self.export_columns:

            self.show_error(
                "Please select at least one data column."
            )

            return

        conn = sqlite3.connect(
            DB_NAME
        )

        cursor = conn.cursor()

        cursor.execute("""
            SELECT
                timestamp,
                systolic,
                diastolic,
                heart_rate,
                stress,
                meditation,
                heart_medication,
                bp_medication,
                heart_med_name,
                bp_med_name,
                daily_note,
                exercise_duration,
                sleep_duration
            FROM health_records
            ORDER BY timestamp DESC
        """)

        records = cursor.fetchall()

        conn.close()

        if not records:

            self.show_error(
                "There is no data available to export."
            )

            return

        # ----------------------------------------------------
        # DATABASE INDEX
        # ----------------------------------------------------

        column_indexes = {
            "Date / Time": 0,
            "Systolic Pressure": 1,
            "Diastolic Pressure": 2,
            "Heart Rate": 3,
            "Stress": 4,
            "Meditation": 5,
            "Heart Medication": 6,
            "BP Medication": 7,
            "Heart Medication Name": 8,
            "BP Medication Name": 9,
            "Daily Note": 10,
            "Exercise Duration": 11,
            "Sleep Duration": 12
        }

        # ----------------------------------------------------
        # PREPARE DATA
        # ----------------------------------------------------

        export_data_rows = []

        for record in records:

            row = []

            for column in self.export_columns:

                index = column_indexes[column]

                value = record[index]

                if column in (
                    "Stress",
                    "Meditation",
                    "Heart Medication",
                    "BP Medication"
                ):

                    value = (
                        "Yes"
                        if value
                        else "No"
                    )

                elif column == "Heart Rate":

                    value = f"{value} BPM"

                elif column == "Exercise Duration":

                    value = f"{value:g} min"

                elif column == "Sleep Duration":

                    value = f"{value:g} hr"

                elif column == "Date / Time":

                    try:

                        date_obj = datetime.strptime(
                            value,
                            "%Y-%m-%d %H:%M:%S"
                        )

                        value = date_obj.strftime(
                            "%d %b %Y  %I:%M %p"
                        )

                    except ValueError:
                        pass

                elif value is None:

                    value = ""

                row.append(
                    value
                )

            export_data_rows.append(
                row
            )

        # ----------------------------------------------------
        # FORMAT
        # ----------------------------------------------------

        export_format = self.export_format_var.get()

        if export_format == "PDF":

            self.export_pdf(
                export_data_rows
            )

        elif export_format == "Excel":

            self.export_excel(
                export_data_rows
            )

    # ========================================================
    # EXPORT EXCEL
    # ========================================================

    def export_excel(
        self,
        data
    ):

        filename = datetime.now().strftime(
            "health_data_%Y%m%d_%H%M%S.xlsx"
        )

        filepath = filedialog.asksaveasfilename(
            title="Export Health Data to Excel",
            defaultextension=".xlsx",
            initialfile=filename,
            filetypes=[
                (
                    "Excel files",
                    "*.xlsx"
                ),
                (
                    "All files",
                    "*.*"
                )
            ]
        )

        if not filepath:
            return

        try:

            workbook = Workbook()

            sheet = workbook.active

            sheet.title = "Heart Health Data"

            # ------------------------------------------------
            # HEADER
            # ------------------------------------------------

            for column_index, column_name in enumerate(
                self.export_columns,
                start=1
            ):

                cell = sheet.cell(
                    row=1,
                    column=column_index,
                    value=column_name
                )

                cell.font = Font(
                    bold=True,
                    color="FFFFFF"
                )

                cell.fill = PatternFill(
                    start_color="0F172A",
                    end_color="0F172A",
                    fill_type="solid"
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center"
                )

            # ------------------------------------------------
            # DATA
            # ------------------------------------------------

            for row_index, row_data in enumerate(
                data,
                start=2
            ):

                for column_index, value in enumerate(
                    row_data,
                    start=1
                ):

                    cell = sheet.cell(
                        row=row_index,
                        column=column_index,
                        value=value
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                        wrap_text=True
                    )

            # ------------------------------------------------
            # WIDTH
            # ------------------------------------------------

            for column_cells in sheet.columns:

                max_length = 0

                column_letter = (
                    column_cells[0].column_letter
                )

                for cell in column_cells:

                    if cell.value is not None:

                        max_length = max(
                            max_length,
                            len(str(cell.value))
                        )

                sheet.column_dimensions[
                    column_letter
                ].width = min(
                    max_length + 3,
                    40
                )

            sheet.freeze_panes = "A2"

            sheet.auto_filter.ref = sheet.dimensions

            workbook.save(
                filepath
            )

            self.show_success(
                "Health data exported successfully to Excel."
            )

        except Exception as error:

            print(
                "Excel export error:",
                error
            )

            self.show_error(
                "Unable to create the Excel file."
            )

    # ========================================================
    # EXPORT PDF
    # ========================================================

    def export_pdf(
        self,
        data
    ):

        filename = datetime.now().strftime(
            "health_data_%Y%m%d_%H%M%S.pdf"
        )

        filepath = filedialog.asksaveasfilename(
            title="Export Health Data to PDF",
            defaultextension=".pdf",
            initialfile=filename,
            filetypes=[
                (
                    "PDF files",
                    "*.pdf"
                ),
                (
                    "All files",
                    "*.*"
                )
            ]
        )

        if not filepath:
            return

        try:

            document = SimpleDocTemplate(
                filepath,
                pagesize=landscape(A4),
                rightMargin=25,
                leftMargin=25,
                topMargin=25,
                bottomMargin=25
            )

            styles = getSampleStyleSheet()

            # =================================================
            # TITLE
            # =================================================

            title_style = styles["Title"]

            title = Paragraph(
                "Heart Health Monitor",
                title_style
            )

            subtitle = Paragraph(
                (
                    "Health Data Export — "
                    f"{datetime.now().strftime('%d %B %Y, %I:%M %p')}"
                ),
                styles["Normal"]
            )

            # =================================================
            # WHITE HEADER STYLE
            # =================================================

            header_style = styles["Normal"].clone(
                "ExportHeaderStyle"
            )

            header_style.fontName = "Helvetica-Bold"
            header_style.fontSize = 7
            header_style.textColor = colors.white
            header_style.alignment = 1
            header_style.leading = 9

            # =================================================
            # DATA STYLE
            # =================================================

            data_style = styles["Normal"].clone(
                "ExportDataStyle"
            )

            data_style.fontName = "Helvetica"
            data_style.fontSize = 7
            data_style.textColor = colors.HexColor(
                "#172033"
            )
            data_style.alignment = 1
            data_style.leading = 9

            # =================================================
            # TABLE DATA
            # =================================================

            table_data = []

            # -------------------------------------------------
            # HEADER
            # -------------------------------------------------

            header_row = []

            for column_name in self.export_columns:

                header_row.append(
                    Paragraph(
                        str(column_name),
                        header_style
                    )
                )

            table_data.append(
                header_row
            )

            # -------------------------------------------------
            # DATA
            # -------------------------------------------------

            for row in data:

                formatted_row = []

                for value in row:

                    formatted_row.append(
                        Paragraph(
                            str(value),
                            data_style
                        )
                    )

                table_data.append(
                    formatted_row
                )

            # =================================================
            # TABLE
            # =================================================

            table = Table(
                table_data,
                repeatRows=1
            )

            table.setStyle(
                TableStyle([

                    # ------------------------------------------------
                    # DARK HEADER
                    # ------------------------------------------------

                    (
                        "BACKGROUND",
                        (0, 0),
                        (-1, 0),
                        colors.HexColor("#0F172A")
                    ),

                    # ------------------------------------------------
                    # WHITE HEADER TEXT
                    # ------------------------------------------------

                    (
                        "TEXTCOLOR",
                        (0, 0),
                        (-1, 0),
                        colors.white
                    ),

                    (
                        "FONTNAME",
                        (0, 0),
                        (-1, 0),
                        "Helvetica-Bold"
                    ),

                    (
                        "FONTSIZE",
                        (0, 0),
                        (-1, 0),
                        7
                    ),

                    # ------------------------------------------------
                    # DATA TEXT
                    # ------------------------------------------------

                    (
                        "TEXTCOLOR",
                        (0, 1),
                        (-1, -1),
                        colors.HexColor("#172033")
                    ),

                    (
                        "FONTNAME",
                        (0, 1),
                        (-1, -1),
                        "Helvetica"
                    ),

                    (
                        "FONTSIZE",
                        (0, 1),
                        (-1, -1),
                        7
                    ),

                    # ------------------------------------------------
                    # ALIGNMENT
                    # ------------------------------------------------

                    (
                        "ALIGN",
                        (0, 0),
                        (-1, -1),
                        "CENTER"
                    ),

                    (
                        "VALIGN",
                        (0, 0),
                        (-1, -1),
                        "MIDDLE"
                    ),

                    # ------------------------------------------------
                    # GRID
                    # ------------------------------------------------

                    (
                        "GRID",
                        (0, 0),
                        (-1, -1),
                        0.5,
                        colors.HexColor("#CBD5E1")
                    ),

                    # ------------------------------------------------
                    # ALTERNATING ROW COLORS
                    # ------------------------------------------------

                    (
                        "ROWBACKGROUNDS",
                        (0, 1),
                        (-1, -1),
                        [
                            colors.white,
                            colors.HexColor("#F8FAFC")
                        ]
                    ),

                    # ------------------------------------------------
                    # PADDING
                    # ------------------------------------------------

                    (
                        "TOPPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "BOTTOMPADDING",
                        (0, 0),
                        (-1, -1),
                        6
                    ),

                    (
                        "LEFTPADDING",
                        (0, 0),
                        (-1, -1),
                        5
                    ),

                    (
                        "RIGHTPADDING",
                        (0, 0),
                        (-1, -1),
                        5
                    )
                ])
            )

            # =================================================
            # BUILD PDF
            # =================================================

            document.build([
                title,
                Spacer(1, 8),
                subtitle,
                Spacer(1, 15),
                table
            ])

            self.show_success(
                "Health data exported successfully to PDF."
            )

        except Exception as error:

            print(
                "PDF export error:",
                error
            )

            self.show_error(
                "Unable to create the PDF file."
            )

    # ========================================================
    # CLEAR FORM
    # ========================================================

    def clear_form(
        self,
        update_status=True
    ):

        self.systolic_var.set("")
        self.diastolic_var.set("")
        self.heart_rate_var.set("")

        self.exercise_var.set("")
        self.sleep_var.set("")

        self.stress_var.set(False)
        self.meditation_var.set(False)

        self.heart_medication_var.set(False)
        self.bp_medication_var.set(False)

        self.heart_med_name_var.set("")
        self.bp_med_name_var.set("")

        self.daily_note_text.delete(
            "1.0",
            "end"
        )

        if update_status:

            self.status_label.configure(
                text="Form cleared.",
                text_color=COLORS["muted"]
            )

    # ========================================================
    # SUCCESS MESSAGE
    # ========================================================

    def show_success(
        self,
        message
    ):

        self.status_label.configure(
            text=f"✓  {message}",
            text_color=COLORS["success_text"]
        )

    # ========================================================
    # ERROR MESSAGE
    # ========================================================

    def show_error(
        self,
        message
    ):

        self.status_label.configure(
            text=f"⚠  {message}",
            text_color=COLORS["error_text"]
        )


# ============================================================
# START APPLICATION
# ============================================================

if __name__ == "__main__":

    create_database()

    app = HealthMonitor()

    app.mainloop()
